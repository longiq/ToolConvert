from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import TableData


_FONT = "Arial"
_THIN = Side(style="thin")
_ALL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FONT    = Font(name=_FONT, bold=True, size=11)
_SUBTITLE_FONT = Font(name=_FONT, size=10)
_HEADER_FONT   = Font(name=_FONT, bold=True, size=10, color="FFFFFF")
_HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
_DATA_FONT     = Font(name=_FONT, size=10)
_ZEBRA_FILLS   = (
    PatternFill("solid", fgColor="FFFFFF"),
    PatternFill("solid", fgColor="EBF3FB"),
)
_SUMMARY_FONT  = Font(name=_FONT, bold=True, size=10)
_SUMMARY_FILL  = PatternFill("solid", fgColor="D9E1F2")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(vertical="center", wrap_text=True)


def _safe_sheet_name(name: str, existing: set[str]) -> str:
    name = name[:31].strip()
    for ch in r'\/*?:[]\x00':
        name = name.replace(ch, "")
    name = name[:31] or "Sheet"
    base = name
    suffix = 2
    while name in existing:
        name = f"{base[:28]} ({suffix})"
        suffix += 1
    existing.add(name)
    return name


def _auto_col_widths(ws, min_width: int = 8, max_width: int = 50):
    for col_cells in ws.columns:
        length = min_width
        for cell in col_cells:
            if cell.value:
                val_len = max(len(str(line)) for line in str(cell.value).splitlines())
                length = max(length, val_len + 2)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length, max_width)


def _write_table_to_sheet(ws, table: TableData, start_row: int = 1, freeze: bool = False) -> int:
    """Write a single TableData to ws starting at start_row. Returns next empty row."""
    row = start_row
    num_cols = max(len(table.headers), max((len(r) for r in table.rows), default=1))
    last_col = get_column_letter(num_cols)

    # Metadata rows
    all_meta = ([table.title] if table.title else []) + table.metadata
    for i, meta_text in enumerate(all_meta):
        cell = ws.cell(row=row, column=1, value=meta_text)
        cell.font = _TITLE_FONT if i == 0 else _SUBTITLE_FONT
        cell.alignment = _CENTER
        if num_cols > 1:
            ws.merge_cells(f"A{row}:{last_col}{row}")
        row += 1

    # Header row
    for col_idx, header in enumerate(table.headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _ALL_BORDER
    row += 1

    # No freeze panes (match Claude)
    ws.freeze_panes = None

    # Data rows
    for row_idx, data_row in enumerate(table.rows):
        is_summary = table.is_summary_row[row_idx] if row_idx < len(table.is_summary_row) else False
        zebra_fill = _ZEBRA_FILLS[row_idx % 2]
        for col_idx in range(1, num_cols + 1):
            val = data_row[col_idx - 1] if col_idx - 1 < len(data_row) else ""
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = _ALL_BORDER
            if is_summary:
                cell.font = _SUMMARY_FONT
                cell.fill = _SUMMARY_FILL
            else:
                cell.font = _DATA_FONT
                cell.fill = zebra_fill
            cell.alignment = _CENTER if col_idx == 1 else _LEFT
        row += 1

    return row


def write_excel(tables: list[TableData], output_path: str) -> str:
    wb = Workbook()
    existing_names: set[str] = set()

    # Sheet 1: "Tổng quan" — all content in order.
    # TEMPORARILY DISABLED per request: summary/overview sheet not needed for now.
    # Source kept (not deleted) so it can be re-enabled later.
    # ws_overview = wb.active
    # ws_overview.title = "Tổng quan"
    # existing_names.add("Tổng quan")
    #
    # current_row = 1
    # for table in tables:
    #     current_row = _write_table_to_sheet(ws_overview, table, start_row=current_row)
    #     current_row += 1  # blank row between tables
    #
    # _auto_col_widths(ws_overview)

    # Remove the default empty sheet created by Workbook() since we skip the overview sheet.
    _default_sheet = wb.active

    # Individual sheets per table
    for table in tables:
        preferred = table.sheet_name or table.title or f"Bảng {len(existing_names)}"
        sheet_name = _safe_sheet_name(preferred, existing_names)
        ws = wb.create_sheet(title=sheet_name)
        _write_table_to_sheet(ws, table, start_row=1)
        _auto_col_widths(ws)
        # Fix STT column width
        if table.headers and table.headers[0].strip().upper() in ("STT", "S.TT", "SỐ TT"):
            ws.column_dimensions["A"].width = 6

    # Drop the leftover default sheet if any real table sheets were added.
    if _default_sheet is not None and len(wb.sheetnames) > 1 and _default_sheet in wb.worksheets:
        wb.remove(_default_sheet)

    wb.save(output_path)
    return output_path
